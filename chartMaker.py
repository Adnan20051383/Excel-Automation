import openpyxl as xl
from openpyxl.chart import DoughnutChart, Reference


def make_chart(file_name):
    work_book = xl.load_workbook(file_name)
    sheet = work_book.active
    chart = DoughnutChart()
    data = Reference(sheet, min_col=4, max_col=4, min_row=2, max_row=sheet.max_row)
    labels = Reference(sheet, min_col=2, max_col=2, min_row=2, max_row=sheet.max_row)
    chart.set_categories(labels)
    chart.add_data(data)
    chart.title = 'Prices'
    chart.style = 26
    sheet.add_chart(chart, "E1")
    work_book.save(file_name)

