import openpyxl as xl


def make_new_prices(file_name, percentage):
    work_book = xl.load_workbook(file_name)
    sheet = work_book.active
    for row in range(2, sheet.max_row + 1):
        current_value = sheet.cell(row, 3).value
        new_value = current_value * percentage
        sheet.cell(row, 4).value = new_value
    work_book.save(file_name)

